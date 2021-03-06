#
# Open Source Voting Results Reporter (ORR) - election results report generator
# Copyright (C) 2018  Chris Jerdonek
#
# This file is part of Open Source Voting Results Reporter (ORR).
#
# ORR is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
#

"""
This module provides support for testing our XLSX-related code.

It exposes functions for reading and introspecting XLSX files.
"""

from pathlib import Path

import openpyxl


# TODO: document / encourage passing a binary stream.
def load(path):
    """
    Return a Workbook object.

    Args:
      path: the path to an xlsx file, as a path-like object.
    """
    wb = openpyxl.load_workbook(filename=path, read_only=True)

    return wb


def get_sheet_names(wb):
    """
    Return the names of the sheets, as a list of strings.

    Args:
      wb: a Workbook object.
    """
    return wb.sheetnames


def get_sheet_rows(worksheet):
    """
    Return the row data as a list of tuples.

    Args:
      wb: an openpyxl.worksheet.read_only.ReadOnlyWorksheet object,
        e.g. as returned from wb.worksheets[0].
    """
    rows = []
    for row in worksheet.iter_rows():
        rows.append(tuple(cell.value for cell in row))

    return rows
